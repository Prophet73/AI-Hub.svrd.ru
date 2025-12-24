"""
Admin API endpoints for Hub.
Includes user management, groups, access control, and admin tools.
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..core.dependencies import get_current_admin
from ..services.audit_service import AuditService
from ..db.base import get_db
from ..models import User, Application, UserGroup, ApplicationAccess, OAuthToken, OAuthCode
from ..models.user_group import user_group_members
from ..models.audit_log import AuditLog
from ..models.login_history import LoginHistory
from ..schemas.admin import (
    UserGroupCreate,
    UserGroupUpdate,
    UserGroupResponse,
    UserGroupListResponse,
    UserGroupMemberInfo,
    AccessGrantRequest,
    AccessRevokeRequest,
    ApplicationAccessResponse,
    AdminStatsResponse,
    UserListResponse,
    UserUpdateRequest,
    BulkUserActionRequest,
    BulkGroupMembershipRequest,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ============ User Management ============

@router.get("/users", response_model=List[UserListResponse])
async def list_users(
    search: Optional[str] = Query(None, description="Search by email or name"),
    is_admin: Optional[bool] = Query(None),
    is_active: Optional[bool] = Query(None),
    group_id: Optional[UUID] = Query(None, description="Filter by group"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all users with filtering and search."""
    query = select(User).options(selectinload(User.groups))

    # Apply filters
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            or_(
                User.email.ilike(search_pattern),
                User.display_name.ilike(search_pattern),
                User.first_name.ilike(search_pattern),
                User.last_name.ilike(search_pattern),
            )
        )

    if is_admin is not None:
        query = query.where(User.is_admin == is_admin)

    if is_active is not None:
        query = query.where(User.is_active == is_active)

    if group_id:
        query = query.join(user_group_members).where(
            user_group_members.c.group_id == group_id
        )

    query = query.order_by(User.created_at.desc()).offset(offset).limit(limit)

    result = await db.execute(query)
    users = result.scalars().all()

    # Build response with group names and app counts
    response = []
    for user in users:
        # Count accessible apps
        access_count = await db.execute(
            select(func.count(ApplicationAccess.id)).where(
                ApplicationAccess.user_id == user.id
            )
        )
        direct_count = access_count.scalar() or 0

        response.append(UserListResponse(
            id=user.id,
            email=user.email,
            display_name=user.display_name,
            first_name=user.first_name,
            last_name=user.last_name,
            middle_name=user.middle_name,
            department=user.department,
            job_title=user.job_title,
            is_active=user.is_active,
            is_admin=user.is_admin,
            last_login_at=user.last_login_at,
            created_at=user.created_at,
            groups=[g.name for g in user.groups],
            app_count=direct_count,
        ))

    return response


@router.patch("/users/{user_id}", response_model=UserListResponse)
async def update_user(
    user_id: UUID,
    data: UserUpdateRequest,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update user properties (active, admin status)."""
    result = await db.execute(
        select(User).options(selectinload(User.groups)).where(User.id == user_id)
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Prevent self-demotion
    if user.id == admin.id and data.is_admin is False:
        raise HTTPException(status_code=400, detail="Cannot remove your own admin status")

    # Track changes for audit
    old_values = {}
    new_values = {}

    if data.is_active is not None and user.is_active != data.is_active:
        old_values["is_active"] = user.is_active
        new_values["is_active"] = data.is_active
        user.is_active = data.is_active

    if data.is_admin is not None and user.is_admin != data.is_admin:
        old_values["is_admin"] = user.is_admin
        new_values["is_admin"] = data.is_admin
        user.is_admin = data.is_admin

    # Log audit event if there were changes
    if old_values:
        await AuditService.log_action(
            db=db,
            user_id=admin.id,
            action="user.update",
            entity_type="user",
            entity_id=user.id,
            old_values=old_values,
            new_values=new_values,
            request=request,
        )

    await db.commit()
    await db.refresh(user)

    return UserListResponse(
        id=user.id,
        email=user.email,
        display_name=user.display_name,
        first_name=user.first_name,
        last_name=user.last_name,
        middle_name=user.middle_name,
        department=user.department,
        job_title=user.job_title,
        is_active=user.is_active,
        is_admin=user.is_admin,
        last_login_at=user.last_login_at,
        created_at=user.created_at,
        groups=[g.name for g in user.groups],
        app_count=0,
    )


@router.post("/users/bulk")
async def bulk_user_action(
    data: BulkUserActionRequest,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Perform bulk actions on users."""
    if admin.id in data.user_ids and data.action in ['deactivate', 'remove_admin']:
        raise HTTPException(status_code=400, detail="Cannot perform this action on yourself")

    result = await db.execute(
        select(User).where(User.id.in_(data.user_ids))
    )
    users = result.scalars().all()

    updated = 0
    updated_user_ids = []
    for user in users:
        if data.action == 'activate':
            if not user.is_active:
                user.is_active = True
                updated += 1
                updated_user_ids.append(str(user.id))
        elif data.action == 'deactivate':
            if user.is_active:
                user.is_active = False
                updated += 1
                updated_user_ids.append(str(user.id))
        elif data.action == 'make_admin':
            if not user.is_admin:
                user.is_admin = True
                updated += 1
                updated_user_ids.append(str(user.id))
        elif data.action == 'remove_admin':
            if user.id != admin.id and user.is_admin:
                user.is_admin = False
                updated += 1
                updated_user_ids.append(str(user.id))

    # Log bulk audit event
    if updated > 0:
        await AuditService.log_action(
            db=db,
            user_id=admin.id,
            action=f"user.bulk_{data.action}",
            entity_type="user",
            entity_id=None,
            old_values=None,
            new_values={"user_ids": updated_user_ids, "count": updated},
            request=request,
        )

    await db.commit()
    return {"updated": updated, "action": data.action}


# ============ Group Management ============

@router.get("/groups", response_model=List[UserGroupListResponse])
async def list_groups(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all user groups."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).order_by(UserGroup.name)
    )
    groups = result.scalars().all()

    return [
        UserGroupListResponse(
            id=g.id,
            name=g.name,
            description=g.description,
            color=g.color,
            member_count=len(g.members),
            created_at=g.created_at,
        )
        for g in groups
    ]


@router.post("/groups", response_model=UserGroupResponse, status_code=201)
async def create_group(
    data: UserGroupCreate,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Create a new user group."""
    # Check name uniqueness
    existing = await db.execute(
        select(UserGroup).where(UserGroup.name == data.name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Group name already exists")

    group = UserGroup(
        name=data.name,
        description=data.description,
        color=data.color,
        created_by=admin.id,
    )
    db.add(group)
    await db.flush()

    # Add initial members
    if data.member_ids:
        users_result = await db.execute(
            select(User).where(User.id.in_(data.member_ids))
        )
        users = users_result.scalars().all()
        group.members = list(users)

    # Log audit event
    await AuditService.log_action(
        db=db,
        user_id=admin.id,
        action="group.create",
        entity_type="group",
        entity_id=group.id,
        old_values=None,
        new_values={"name": group.name, "member_count": len(data.member_ids or [])},
        request=request,
    )

    await db.commit()
    await db.refresh(group)

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.get("/groups/{group_id}", response_model=UserGroupResponse)
async def get_group(
    group_id: UUID,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get group details with members."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.patch("/groups/{group_id}", response_model=UserGroupResponse)
async def update_group(
    group_id: UUID,
    data: UserGroupUpdate,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update group properties."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if data.name is not None:
        # Check uniqueness
        existing = await db.execute(
            select(UserGroup).where(UserGroup.name == data.name, UserGroup.id != group_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Group name already exists")
        group.name = data.name

    if data.description is not None:
        group.description = data.description
    if data.color is not None:
        group.color = data.color

    await db.commit()
    await db.refresh(group)

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.delete("/groups/{group_id}", status_code=204)
async def delete_group(
    group_id: UUID,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Delete a group."""
    result = await db.execute(
        select(UserGroup).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Log audit event before deletion
    await AuditService.log_action(
        db=db,
        user_id=admin.id,
        action="group.delete",
        entity_type="group",
        entity_id=group.id,
        old_values={"name": group.name},
        new_values=None,
        request=request,
    )

    await db.delete(group)
    await db.commit()


@router.post("/groups/{group_id}/members")
async def add_group_members(
    group_id: UUID,
    data: BulkGroupMembershipRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Add or remove members from a group."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    users_result = await db.execute(
        select(User).where(User.id.in_(data.user_ids))
    )
    users = users_result.scalars().all()

    current_member_ids = {m.id for m in group.members}
    updated = 0

    if data.action == 'add':
        for user in users:
            if user.id not in current_member_ids:
                group.members.append(user)
                updated += 1
    elif data.action == 'remove':
        group.members = [m for m in group.members if m.id not in data.user_ids]
        updated = len(data.user_ids)

    await db.commit()
    return {"updated": updated, "action": data.action}


# ============ Access Control ============

@router.get("/applications/{app_id}/access", response_model=ApplicationAccessResponse)
async def get_application_access(
    app_id: UUID,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get access configuration for an application."""
    result = await db.execute(
        select(Application).where(Application.id == app_id)
    )
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Get direct user access
    direct_access = await db.execute(
        select(ApplicationAccess).options(selectinload(ApplicationAccess.user)).where(
            ApplicationAccess.application_id == app_id,
            ApplicationAccess.user_id.isnot(None)
        )
    )
    direct_users = [
        UserGroupMemberInfo(
            id=a.user.id,
            email=a.user.email,
            display_name=a.user.display_name,
            first_name=a.user.first_name,
            last_name=a.user.last_name,
            department=a.user.department,
        )
        for a in direct_access.scalars().all()
    ]

    # Get group access
    group_access = await db.execute(
        select(ApplicationAccess).options(
            selectinload(ApplicationAccess.group).selectinload(UserGroup.members)
        ).where(
            ApplicationAccess.application_id == app_id,
            ApplicationAccess.group_id.isnot(None)
        )
    )
    groups = [
        UserGroupListResponse(
            id=a.group.id,
            name=a.group.name,
            description=a.group.description,
            color=a.group.color,
            member_count=len(a.group.members),
            created_at=a.group.created_at,
        )
        for a in group_access.scalars().all()
    ]

    return ApplicationAccessResponse(
        application_id=app.id,
        application_name=app.name,
        is_public=app.is_public or False,
        direct_users=direct_users,
        groups=groups,
    )


@router.post("/access/grant")
async def grant_access(
    data: AccessGrantRequest,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Grant access to an application for users or groups."""
    # Verify application exists
    app_result = await db.execute(
        select(Application).where(Application.id == data.application_id)
    )
    app = app_result.scalar_one_or_none()
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    granted = 0

    # Grant to users
    for user_id in data.user_ids:
        # Check if already exists
        existing = await db.execute(
            select(ApplicationAccess).where(
                ApplicationAccess.user_id == user_id,
                ApplicationAccess.application_id == data.application_id
            )
        )
        if not existing.scalar_one_or_none():
            access = ApplicationAccess(
                user_id=user_id,
                application_id=data.application_id,
                granted_by=admin.id,
            )
            db.add(access)
            granted += 1

    # Grant to groups
    for group_id in data.group_ids:
        existing = await db.execute(
            select(ApplicationAccess).where(
                ApplicationAccess.group_id == group_id,
                ApplicationAccess.application_id == data.application_id
            )
        )
        if not existing.scalar_one_or_none():
            access = ApplicationAccess(
                group_id=group_id,
                application_id=data.application_id,
                granted_by=admin.id,
            )
            db.add(access)
            granted += 1

    # Log audit event
    if granted > 0:
        await AuditService.log_action(
            db=db,
            user_id=admin.id,
            action="access.grant",
            entity_type="application",
            entity_id=data.application_id,
            old_values=None,
            new_values={
                "application": app.name,
                "user_ids": [str(u) for u in data.user_ids],
                "group_ids": [str(g) for g in data.group_ids],
                "granted_count": granted,
            },
            request=request,
        )

    await db.commit()
    return {"granted": granted}


@router.post("/access/revoke")
async def revoke_access(
    data: AccessRevokeRequest,
    request: Request,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Revoke access to an application from users or groups."""
    # Get app name for audit log
    app_result = await db.execute(
        select(Application).where(Application.id == data.application_id)
    )
    app = app_result.scalar_one_or_none()

    revoked = 0

    # Revoke from users
    if data.user_ids:
        result = await db.execute(
            delete(ApplicationAccess).where(
                ApplicationAccess.user_id.in_(data.user_ids),
                ApplicationAccess.application_id == data.application_id
            )
        )
        revoked += result.rowcount

    # Revoke from groups
    if data.group_ids:
        result = await db.execute(
            delete(ApplicationAccess).where(
                ApplicationAccess.group_id.in_(data.group_ids),
                ApplicationAccess.application_id == data.application_id
            )
        )
        revoked += result.rowcount

    # Log audit event
    if revoked > 0:
        await AuditService.log_action(
            db=db,
            user_id=admin.id,
            action="access.revoke",
            entity_type="application",
            entity_id=data.application_id,
            old_values={
                "user_ids": [str(u) for u in (data.user_ids or [])],
                "group_ids": [str(g) for g in (data.group_ids or [])],
            },
            new_values={"revoked_count": revoked},
            request=request,
        )

    await db.commit()
    return {"revoked": revoked}


@router.patch("/applications/{app_id}/public")
async def set_application_public(
    app_id: UUID,
    is_public: bool = Query(...),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Set whether an application is public (visible to all users)."""
    result = await db.execute(
        select(Application).where(Application.id == app_id)
    )
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    app.is_public = is_public
    await db.commit()

    return {"application_id": app_id, "is_public": is_public}


# ============ Admin Stats & Tools ============

@router.get("/stats", response_model=AdminStatsResponse)
async def get_admin_stats(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get dashboard statistics."""
    now = datetime.now(timezone.utc)

    # User stats
    users_total = await db.execute(select(func.count(User.id)))
    users_active = await db.execute(select(func.count(User.id)).where(User.is_active == True))
    users_admin = await db.execute(select(func.count(User.id)).where(User.is_admin == True))

    # App stats
    apps_total = await db.execute(select(func.count(Application.id)))
    apps_active = await db.execute(select(func.count(Application.id)).where(Application.is_active == True))

    # Group stats
    groups_total = await db.execute(select(func.count(UserGroup.id)))

    # Token stats
    tokens_total = await db.execute(select(func.count(OAuthToken.id)))
    tokens_active = await db.execute(
        select(func.count(OAuthToken.id)).where(
            OAuthToken.expires_at > now,
            OAuthToken.revoked_at.is_(None)
        )
    )

    return AdminStatsResponse(
        users={
            "total": users_total.scalar() or 0,
            "active": users_active.scalar() or 0,
            "admins": users_admin.scalar() or 0,
        },
        applications={
            "total": apps_total.scalar() or 0,
            "active": apps_active.scalar() or 0,
        },
        groups={
            "total": groups_total.scalar() or 0,
        },
        tokens={
            "total": tokens_total.scalar() or 0,
            "active": tokens_active.scalar() or 0,
        },
        database={
            "status": "healthy",
        },
        generated_at=now,
    )


@router.get("/health")
async def get_service_health(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Check health status of all connected services."""
    import time
    import httpx
    from ..core.config import settings

    services = []
    now = datetime.now(timezone.utc)

    # 1. Backend API health (self)
    services.append({
        "name": "Backend API",
        "status": "healthy",
        "latency_ms": 0,
        "last_check": now.isoformat(),
        "message": "Работает нормально",
    })

    # 2. PostgreSQL database health
    try:
        start_time = time.time()
        await db.execute(select(func.count(User.id)))
        latency = int((time.time() - start_time) * 1000)
        services.append({
            "name": "PostgreSQL",
            "status": "healthy",
            "latency_ms": latency,
            "last_check": now.isoformat(),
            "message": "Подключение активно",
        })
    except Exception as e:
        services.append({
            "name": "PostgreSQL",
            "status": "unhealthy",
            "latency_ms": None,
            "last_check": now.isoformat(),
            "message": f"Ошибка: {str(e)}",
        })

    # 3. Gemini AI health
    try:
        start_time = time.time()
        gemini_api_key = getattr(settings, 'GEMINI_API_KEY', None)
        if gemini_api_key:
            async with httpx.AsyncClient(timeout=5.0) as client:
                # Just check if we can reach the API endpoint
                response = await client.get(
                    f"https://generativelanguage.googleapis.com/v1beta/models?key={gemini_api_key}"
                )
                latency = int((time.time() - start_time) * 1000)
                if response.status_code == 200:
                    services.append({
                        "name": "Gemini AI",
                        "status": "healthy",
                        "latency_ms": latency,
                        "last_check": now.isoformat(),
                        "message": "API доступен",
                    })
                else:
                    services.append({
                        "name": "Gemini AI",
                        "status": "degraded",
                        "latency_ms": latency,
                        "last_check": now.isoformat(),
                        "message": f"Код ответа: {response.status_code}",
                    })
        else:
            services.append({
                "name": "Gemini AI",
                "status": "degraded",
                "latency_ms": None,
                "last_check": now.isoformat(),
                "message": "API ключ не настроен",
            })
    except httpx.TimeoutException:
        services.append({
            "name": "Gemini AI",
            "status": "degraded",
            "latency_ms": 5000,
            "last_check": now.isoformat(),
            "message": "Таймаут подключения",
        })
    except Exception as e:
        services.append({
            "name": "Gemini AI",
            "status": "unhealthy",
            "latency_ms": None,
            "last_check": now.isoformat(),
            "message": f"Ошибка: {str(e)[:50]}",
        })

    # 4. ADFS/SSO health
    try:
        start_time = time.time()
        oidc_url = getattr(settings, 'OIDC_DISCOVERY_URL', None)
        if oidc_url:
            async with httpx.AsyncClient(timeout=5.0) as client:
                response = await client.get(oidc_url)
                latency = int((time.time() - start_time) * 1000)
                if response.status_code == 200:
                    services.append({
                        "name": "ADFS SSO",
                        "status": "healthy",
                        "latency_ms": latency,
                        "last_check": now.isoformat(),
                        "message": "Сервер авторизации доступен",
                    })
                else:
                    services.append({
                        "name": "ADFS SSO",
                        "status": "degraded",
                        "latency_ms": latency,
                        "last_check": now.isoformat(),
                        "message": f"Код ответа: {response.status_code}",
                    })
        else:
            services.append({
                "name": "ADFS SSO",
                "status": "degraded",
                "latency_ms": None,
                "last_check": now.isoformat(),
                "message": "URL не настроен",
            })
    except httpx.TimeoutException:
        services.append({
            "name": "ADFS SSO",
            "status": "unhealthy",
            "latency_ms": 5000,
            "last_check": now.isoformat(),
            "message": "Таймаут подключения",
        })
    except Exception as e:
        services.append({
            "name": "ADFS SSO",
            "status": "unhealthy",
            "latency_ms": None,
            "last_check": now.isoformat(),
            "message": f"Ошибка: {str(e)[:50]}",
        })

    # Calculate overall status
    statuses = [s["status"] for s in services]
    if all(s == "healthy" for s in statuses):
        overall_status = "healthy"
    elif "unhealthy" in statuses:
        overall_status = "unhealthy"
    else:
        overall_status = "degraded"

    return {
        "status": overall_status,
        "services": services,
        "checked_at": now.isoformat(),
    }


@router.post("/cleanup-tokens")
async def cleanup_tokens(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Clean up expired OAuth codes and tokens."""
    now = datetime.now(timezone.utc)

    # Delete expired codes
    codes_result = await db.execute(
        delete(OAuthCode).where(OAuthCode.expires_at < now)
    )

    # Delete expired tokens
    tokens_result = await db.execute(
        delete(OAuthToken).where(OAuthToken.expires_at < now)
    )

    await db.commit()

    return {
        "deleted_codes": codes_result.rowcount,
        "deleted_tokens": tokens_result.rowcount,
    }


@router.get("/export/users")
async def export_users_excel(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export all users to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    result = await db.execute(
        select(User).options(selectinload(User.groups)).order_by(User.created_at.desc())
    )
    users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Headers
    headers = ['Email', 'Display Name', 'First Name', 'Last Name', 'Middle Name',
               'Department', 'Job Title', 'Groups', 'Admin', 'Active', 'Last Login', 'Created']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.email)
        ws.cell(row=row, column=2, value=user.display_name)
        ws.cell(row=row, column=3, value=user.first_name)
        ws.cell(row=row, column=4, value=user.last_name)
        ws.cell(row=row, column=5, value=user.middle_name)
        ws.cell(row=row, column=6, value=user.department)
        ws.cell(row=row, column=7, value=user.job_title)
        ws.cell(row=row, column=8, value=', '.join(g.name for g in user.groups))
        ws.cell(row=row, column=9, value='Yes' if user.is_admin else 'No')
        ws.cell(row=row, column=10, value='Yes' if user.is_active else 'No')
        ws.cell(row=row, column=11, value=user.last_login_at.strftime('%Y-%m-%d %H:%M') if user.last_login_at else '')
        ws.cell(row=row, column=12, value=user.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-width
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"hub_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/applications")
async def export_applications_excel(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export all applications to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    result = await db.execute(
        select(Application).order_by(Application.name)
    )
    apps = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    headers = ['Name', 'Slug', 'Client ID', 'Base URL', 'Active', 'Public', 'Created']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for row, app in enumerate(apps, 2):
        ws.cell(row=row, column=1, value=app.name)
        ws.cell(row=row, column=2, value=app.slug)
        ws.cell(row=row, column=3, value=app.client_id)
        ws.cell(row=row, column=4, value=app.base_url)
        ws.cell(row=row, column=5, value='Yes' if app.is_active else 'No')
        ws.cell(row=row, column=6, value='Yes' if app.is_public else 'No')
        ws.cell(row=row, column=7, value=app.created_at.strftime('%Y-%m-%d %H:%M'))

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"hub_applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ Applications Management ============

@router.get("/applications")
async def list_applications_admin(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all applications with full details (admin only)."""
    result = await db.execute(
        select(Application).order_by(Application.name)
    )
    apps = result.scalars().all()

    return [
        {
            "id": str(app.id),
            "name": app.name,
            "slug": app.slug,
            "description": app.description,
            "base_url": app.base_url,
            "icon_url": app.icon_url,
            "client_id": app.client_id,
            "redirect_uris": app.redirect_uris or [],
            "is_active": app.is_active,
            "is_public": app.is_public or False,
            "allowed_departments": app.allowed_departments or [],
            "created_at": app.created_at.isoformat() if app.created_at else None,
        }
        for app in apps
    ]


@router.get("/departments")
async def list_departments(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get all unique departments from users (for autocomplete)."""
    result = await db.execute(
        select(User.department).where(User.department.isnot(None)).distinct()
    )
    departments = [row[0] for row in result.fetchall() if row[0]]
    return sorted(departments)


@router.patch("/applications/{app_id}/departments")
async def set_application_departments(
    app_id: UUID,
    departments: List[str],
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Set allowed departments for an application. Empty list = available to all."""
    result = await db.execute(
        select(Application).where(Application.id == app_id)
    )
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    app.allowed_departments = departments
    await db.commit()

    return {
        "application_id": str(app_id),
        "allowed_departments": departments,
        "message": "Доступно всем" if not departments else f"Доступно: {', '.join(departments)}"
    }


# ============ Audit Log & Login History ============

@router.get("/audit-log")
async def get_audit_log(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=10, le=100),
    action: Optional[str] = Query(None, description="Filter by action type"),
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    user_id: Optional[UUID] = Query(None, description="Filter by user who performed action"),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get paginated audit log with filters."""
    query = select(AuditLog).options(selectinload(AuditLog.user))

    # Apply filters
    if action:
        query = query.where(AuditLog.action.ilike(f"%{action}%"))
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    if user_id:
        query = query.where(AuditLog.user_id == user_id)

    # Get total count
    count_query = select(func.count(AuditLog.id))
    if action:
        count_query = count_query.where(AuditLog.action.ilike(f"%{action}%"))
    if entity_type:
        count_query = count_query.where(AuditLog.entity_type == entity_type)
    if user_id:
        count_query = count_query.where(AuditLog.user_id == user_id)

    total = (await db.execute(count_query)).scalar() or 0

    # Apply pagination
    offset = (page - 1) * per_page
    query = query.order_by(AuditLog.created_at.desc()).offset(offset).limit(per_page)

    result = await db.execute(query)
    logs = result.scalars().all()

    return {
        "items": [
            {
                "id": str(log.id),
                "user_id": str(log.user_id) if log.user_id else None,
                "user_email": log.user.email if log.user else None,
                "user_name": log.user.display_name if log.user else None,
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": str(log.entity_id) if log.entity_id else None,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.get("/login-history")
async def get_login_history(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=10, le=100),
    user_id: Optional[UUID] = Query(None, description="Filter by user"),
    login_type: Optional[str] = Query(None, description="Filter by login type (sso, dev)"),
    success: Optional[bool] = Query(None, description="Filter by success status"),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get paginated login history with filters."""
    query = select(LoginHistory).options(selectinload(LoginHistory.user))

    # Apply filters
    if user_id:
        query = query.where(LoginHistory.user_id == user_id)
    if login_type:
        query = query.where(LoginHistory.login_type == login_type)
    if success is not None:
        query = query.where(LoginHistory.success == success)

    # Get total count
    count_query = select(func.count(LoginHistory.id))
    if user_id:
        count_query = count_query.where(LoginHistory.user_id == user_id)
    if login_type:
        count_query = count_query.where(LoginHistory.login_type == login_type)
    if success is not None:
        count_query = count_query.where(LoginHistory.success == success)

    total = (await db.execute(count_query)).scalar() or 0

    # Apply pagination
    offset = (page - 1) * per_page
    query = query.order_by(LoginHistory.created_at.desc()).offset(offset).limit(per_page)

    result = await db.execute(query)
    logs = result.scalars().all()

    return {
        "items": [
            {
                "id": str(log.id),
                "user_id": str(log.user_id) if log.user_id else None,
                "user_email": log.user.email if log.user else None,
                "user_name": log.user.display_name if log.user else None,
                "login_type": log.login_type,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "success": log.success,
                "failure_reason": log.failure_reason,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.get("/stats/logins")
async def get_login_stats(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get login statistics."""
    from datetime import timedelta

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    # Count logins
    today_count = await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.created_at >= today_start,
            LoginHistory.success == True
        )
    )
    week_count = await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.created_at >= week_ago,
            LoginHistory.success == True
        )
    )
    month_count = await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.created_at >= month_ago,
            LoginHistory.success == True
        )
    )

    return {
        "today": today_count.scalar() or 0,
        "last_7_days": week_count.scalar() or 0,
        "last_30_days": month_count.scalar() or 0,
    }


@router.get("/stats/usage")
async def get_usage_stats(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get comprehensive usage statistics for the dashboard."""
    from datetime import timedelta
    from ..models.chat import ChatMessage, ChatUsage, MODEL_PRICING, calculate_cost
    from ..models.prompt import Prompt

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    # Chat statistics
    total_messages = (await db.execute(select(func.count(ChatMessage.id)))).scalar() or 0
    messages_today = (await db.execute(
        select(func.count(ChatMessage.id)).where(ChatMessage.created_at >= today_start)
    )).scalar() or 0
    messages_week = (await db.execute(
        select(func.count(ChatMessage.id)).where(ChatMessage.created_at >= week_ago)
    )).scalar() or 0

    # Messages by day (last 7 days)
    messages_by_day_query = await db.execute(
        select(
            func.date(ChatMessage.created_at).label('date'),
            func.count(ChatMessage.id).label('count')
        )
        .where(ChatMessage.created_at >= week_ago)
        .group_by(func.date(ChatMessage.created_at))
        .order_by(func.date(ChatMessage.created_at))
    )
    messages_by_day = [{"date": str(row.date), "count": row.count} for row in messages_by_day_query.all()]

    # Active users (users who sent messages)
    active_users_today = (await db.execute(
        select(func.count(func.distinct(ChatMessage.user_id)))
        .where(ChatMessage.created_at >= today_start)
    )).scalar() or 0
    active_users_week = (await db.execute(
        select(func.count(func.distinct(ChatMessage.user_id)))
        .where(ChatMessage.created_at >= week_ago)
    )).scalar() or 0

    # Top users by message count
    top_users_query = await db.execute(
        select(
            User.display_name,
            User.email,
            func.count(ChatMessage.id).label('message_count')
        )
        .join(User, ChatMessage.user_id == User.id)
        .where(ChatMessage.created_at >= week_ago)
        .group_by(User.id, User.display_name, User.email)
        .order_by(func.count(ChatMessage.id).desc())
        .limit(10)
    )
    top_users = [
        {"name": row.display_name or row.email, "messages": row.message_count}
        for row in top_users_query.all()
    ]

    # Chat sessions count
    total_sessions = (await db.execute(
        select(func.count(func.distinct(ChatMessage.session_id)))
    )).scalar() or 0
    sessions_today = (await db.execute(
        select(func.count(func.distinct(ChatMessage.session_id)))
        .where(ChatMessage.created_at >= today_start)
    )).scalar() or 0

    # Prompt statistics
    total_prompts = (await db.execute(
        select(func.count(Prompt.id)).where(Prompt.is_active == True)
    )).scalar() or 0

    # Top prompts by usage
    top_prompts_query = await db.execute(
        select(Prompt.name, Prompt.category, Prompt.usage_count)
        .where(Prompt.is_active == True, Prompt.usage_count > 0)
        .order_by(Prompt.usage_count.desc())
        .limit(10)
    )
    top_prompts = [
        {"name": row.name, "category": row.category, "usage_count": row.usage_count}
        for row in top_prompts_query.all()
    ]

    # User statistics
    total_users = (await db.execute(select(func.count(User.id)))).scalar() or 0
    active_users = (await db.execute(
        select(func.count(User.id)).where(User.is_active == True)
    )).scalar() or 0

    # Login statistics
    logins_today = (await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.created_at >= today_start,
            LoginHistory.success == True
        )
    )).scalar() or 0
    logins_week = (await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.created_at >= week_ago,
            LoginHistory.success == True
        )
    )).scalar() or 0

    # Token usage and costs
    tokens_today_query = await db.execute(
        select(
            func.coalesce(func.sum(ChatMessage.input_tokens), 0).label('input'),
            func.coalesce(func.sum(ChatMessage.output_tokens), 0).label('output')
        )
        .where(ChatMessage.created_at >= today_start, ChatMessage.role == 'assistant')
    )
    tokens_today = tokens_today_query.one()

    tokens_week_query = await db.execute(
        select(
            func.coalesce(func.sum(ChatMessage.input_tokens), 0).label('input'),
            func.coalesce(func.sum(ChatMessage.output_tokens), 0).label('output')
        )
        .where(ChatMessage.created_at >= week_ago, ChatMessage.role == 'assistant')
    )
    tokens_week = tokens_week_query.one()

    tokens_total_query = await db.execute(
        select(
            func.coalesce(func.sum(ChatMessage.input_tokens), 0).label('input'),
            func.coalesce(func.sum(ChatMessage.output_tokens), 0).label('output')
        )
        .where(ChatMessage.role == 'assistant')
    )
    tokens_total = tokens_total_query.one()

    # Calculate costs using default model pricing (gemini-2.0-flash)
    cost_today = calculate_cost("gemini-2.0-flash", tokens_today.input, tokens_today.output)
    cost_week = calculate_cost("gemini-2.0-flash", tokens_week.input, tokens_week.output)
    cost_total = calculate_cost("gemini-2.0-flash", tokens_total.input, tokens_total.output)

    # Cost by day (last 7 days)
    costs_by_day_query = await db.execute(
        select(
            func.date(ChatMessage.created_at).label('date'),
            func.coalesce(func.sum(ChatMessage.input_tokens), 0).label('input_tokens'),
            func.coalesce(func.sum(ChatMessage.output_tokens), 0).label('output_tokens')
        )
        .where(ChatMessage.created_at >= week_ago, ChatMessage.role == 'assistant')
        .group_by(func.date(ChatMessage.created_at))
        .order_by(func.date(ChatMessage.created_at))
    )
    costs_by_day = [
        {
            "date": str(row.date),
            "input_tokens": row.input_tokens,
            "output_tokens": row.output_tokens,
            "cost": calculate_cost("gemini-2.0-flash", row.input_tokens, row.output_tokens)
        }
        for row in costs_by_day_query.all()
    ]

    return {
        "chat": {
            "total_messages": total_messages,
            "messages_today": messages_today,
            "messages_week": messages_week,
            "messages_by_day": messages_by_day,
            "total_sessions": total_sessions,
            "sessions_today": sessions_today,
            "active_users_today": active_users_today,
            "active_users_week": active_users_week,
            "top_users": top_users,
        },
        "prompts": {
            "total_prompts": total_prompts,
            "top_prompts": top_prompts,
        },
        "users": {
            "total_users": total_users,
            "active_users": active_users,
        },
        "logins": {
            "logins_today": logins_today,
            "logins_week": logins_week,
        },
        "costs": {
            "tokens_today": {"input": tokens_today.input, "output": tokens_today.output},
            "tokens_week": {"input": tokens_week.input, "output": tokens_week.output},
            "tokens_total": {"input": tokens_total.input, "output": tokens_total.output},
            "cost_today_usd": round(cost_today, 4),
            "cost_week_usd": round(cost_week, 4),
            "cost_total_usd": round(cost_total, 4),
            "costs_by_day": costs_by_day,
            "pricing": MODEL_PRICING,
        },
    }
